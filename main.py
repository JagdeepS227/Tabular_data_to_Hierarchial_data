import openpyxl 
import json

EMP_INDEX = 2
DES_INDEX = 3
DEP_INDEX = 4
NAME_INDEX = 5
MAN_INDEX = 6

class Employee:
    def __init__(self, name, employee_id, manager_employee_id, department, designation, list_of_subordinate, dict_subordinate):
        self.name = name
        self.employee_id = employee_id
        self.manager_employee_id = manager_employee_id
        self.department = department
        self.designation = designation        
        self.list_of_subordinate = list_of_subordinate
        self.dict_subordinate = dict_subordinate

    def add_subordinate(self, person, index=0):
        self.list_of_subordinate.insert(index, person)

   
    dict_subordinate = {}
    name = ''
    employee_id = ''
    manager_employee_id = ''
    department = ''
    designation = ''
    list_of_subordinate = []


def make_dictionary(emp):
    emp.dict_subordinate['name'] = emp.name
    emp.dict_subordinate['id'] = emp.employee_id
    emp.dict_subordinate['reportees'] = []

    #print(emp.name, len(emp.list_of_subordinate))
    if(emp == None or len(emp.list_of_subordinate) == 0):
        if(emp.dict_subordinate == None):
            return {}
        return emp.dict_subordinate
        
    for i in emp.list_of_subordinate:
        if(emp.dict_subordinate == None):
            return ''
        emp.dict_subordinate['reportees'].insert(0, make_dictionary(i))
    return emp.dict_subordinate
    



path = input()
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
dict_objects = {}
ceo = None

for k in range(2, sheet_obj.max_row+1):
    new_emp = Employee(sheet_obj.cell(row=k, column=NAME_INDEX).value,
            sheet_obj.cell(row=k, column=EMP_INDEX).value,
            sheet_obj.cell(row=k, column=MAN_INDEX).value,
            sheet_obj.cell(row=k, column=DEP_INDEX).value,
            sheet_obj.cell(row=k, column=DES_INDEX).value,   
            [],
            {},             
    )
    dict_objects[new_emp.employee_id] = new_emp
    if(sheet_obj.cell(row=k, column=MAN_INDEX).value == None):
        ceo = new_emp

for emp in dict_objects.values():
    if(emp.manager_employee_id == None):
        continue    
    manager_key = emp.manager_employee_id
    if(manager_key not in dict_objects.keys()):
        manager_key = ceo.employee_id
    manager = dict_objects[str(manager_key)]
    manager.add_subordinate(emp)


make_dictionary(ceo)


print(ceo.dict_subordinate)
print(" ")
print(json.dumps(ceo.dict_subordinate,sort_keys=True, indent=4))
